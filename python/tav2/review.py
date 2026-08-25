"""可选审校模式：翻译结果写审校表 xlsx（含说话人/上下文），人工确认后回填。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from tav2.models import Document


def write_review_sheet(
    project_dir: Path,
    document: Document,
    translations: dict[str, str],
) -> Path:
    """把本次译文写入审校表 xlsx。返回文件路径。"""

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "审校"
    ws.append(
        [
            "类型",
            "文件",
            "标识符",
            "序号",
            "源文本",
            "说话人",
            "上下文",
            "机器译文",
            "人工译文",
            "状态",
        ]
    )
    for scene in document.scenes:
        for idx, unit in enumerate(scene.units):
            text = translations.get(unit.unit_id)
            if text is None:
                continue
            extra = unit.extra
            context = _context_for(scene.units, idx)
            if unit.kind == "string":
                ws.append(
                    [
                        "string",
                        extra.get("file", ""),
                        unit.source,
                        "",
                        unit.source,
                        "",
                        context,
                        text,
                        "",
                        "待审",
                    ]
                )
            else:
                ws.append(
                    [
                        "dialogue",
                        extra.get("file", ""),
                        extra.get("identifier", ""),
                        extra.get("say_index", 0),
                        unit.source,
                        unit.speaker,
                        context,
                        text,
                        "",
                        "待审",
                    ]
                )
    project_dir.mkdir(parents=True, exist_ok=True)
    import time

    path = project_dir / f"review_{document.lang}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path


def _context_for(units: list[Any], idx: int, max_chars: int = 240) -> str:
    """取场景内前后各 2 句原文作为上下文，方便人工判断。"""

    parts: list[str] = []
    for j in range(max(0, idx - 2), idx):
        if units[j].source.strip():
            parts.append(f"前: {units[j].source[:60]}")
    for j in range(idx + 1, min(len(units), idx + 3)):
        if units[j].source.strip():
            parts.append(f"后: {units[j].source[:60]}")
    return " | ".join(parts)[:max_chars]


def read_review_sheet(path: Path) -> list[dict[str, Any]]:
    """读取审校表为行字典列表。"""

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    header: list[str] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0:
            header = [str(c or "").strip() for c in row]
            continue
        if not row or not any(row):
            continue
        rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    wb.close()
    return rows
